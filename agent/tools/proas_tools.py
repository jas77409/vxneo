"""
PROAS Management System Integration
Reads/writes to PROAS_Management_System.xlsm on Google Drive
File ID: 1BjY8fdEF4Tsu_PcefqIX5OZ7mg3QsXh-
"""
import json, io, os
from datetime import datetime
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import openpyxl

PROAS_FILE_ID = os.environ.get('PROAS_FILE_ID', '')  # set in .env
TOKEN_PATH = '/root/companion/google_token.json'

def _drive():
    with open(TOKEN_PATH) as f:
        token = json.load(f)
    creds = Credentials.from_authorized_user_info(token)
    return build('drive', 'v3', credentials=creds)

def _download_wb():
    drive = _drive()
    request = drive.files().get_media(fileId=PROAS_FILE_ID)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True), buf

def _upload_wb(wb):
    drive = _drive()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    media = MediaIoBaseUpload(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    drive.files().update(fileId=PROAS_FILE_ID, media_body=media).execute()
    return True

def get_clients():
    """Get all active PROAS clients"""
    try:
        wb, _ = _download_wb()
        ws = wb['Clients']
        clients = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            if headers is None:
                headers = row
                continue
            client = dict(zip(headers, row))
            if client.get('Active') == 'Yes':
                clients.append(client)
        return {'clients': clients, 'count': len(clients)}
    except Exception as e:
        return {'error': str(e)}

def get_employees():
    """Get all PROAS employees"""
    try:
        wb, _ = _download_wb()
        ws = wb['Employees']
        employees = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            if headers is None:
                headers = row
                continue
            emp = dict(zip(headers, row))
            if emp.get('Employee_ID'):
                employees.append(emp)
        return {'employees': employees, 'count': len(employees)}
    except Exception as e:
        return {'error': str(e)}

def get_timesheets(month=None, client_id=None, employee_id=None):
    """Get timesheets, optionally filtered by month, client, or employee"""
    try:
        wb, _ = _download_wb()
        ws = wb['Timesheets']
        entries = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            if headers is None:
                headers = row
                continue
            entry = dict(zip(headers, row))
            if not entry.get('Date'):
                continue
            if month and hasattr(entry['Date'], 'month'):
                if entry['Date'].month != month:
                    continue
            if client_id and entry.get('Client_ID') != client_id:
                continue
            if employee_id and entry.get('Employee_ID') != employee_id:
                continue
            entries.append(entry)
        return {'timesheets': entries, 'count': len(entries)}
    except Exception as e:
        return {'error': str(e)}

def add_timesheet_entry(date, employee_id, client_id, service_type, hours, notes=''):
    """Add a new timesheet entry"""
    try:
        wb, _ = _download_wb()

        # Get employee rate
        emp_data = get_employees()
        emp_rate = 0
        for e in emp_data.get('employees', []):
            if e['Employee_ID'] == employee_id:
                emp_rate = e.get('Hourly_Pay_Rate', 0) or 0
                break

        # Get client rate
        cli_data = get_clients()
        cli_rate = 0
        for c in cli_data.get('clients', []):
            if c['Client_ID'] == client_id:
                cli_rate = c.get('Hourly_Charge_Rate_1', 0) or 0
                break

        revenue = hours * cli_rate
        emp_cost = hours * emp_rate
        gross_profit = revenue - emp_cost
        admin_overhead = hours * 0.1
        net_profit = gross_profit - admin_overhead

        if isinstance(date, str):
            date = datetime.strptime(date, '%Y-%m-%d')

        ws = wb['Timesheets']
        ws.append([
            date, employee_id, client_id, service_type,
            hours, emp_rate, cli_rate, revenue,
            emp_cost, gross_profit, notes,
            admin_overhead, net_profit, cli_rate, 'Rate 1'
        ])

        _upload_wb(wb)
        return {
            'success': True,
            'entry': {
                'date': str(date.date()),
                'employee': employee_id,
                'client': client_id,
                'hours': hours,
                'revenue': revenue,
                'net_profit': net_profit
            }
        }
    except Exception as e:
        return {'error': str(e)}

def get_monthly_summary(month=None, year=None):
    """Get monthly billing summary per client"""
    try:
        now = datetime.now()
        month = month or now.month
        year = year or now.year

        ts = get_timesheets(month=month)
        clients = {c['Client_ID']: c for c in get_clients().get('clients', [])}

        summary = {}
        for entry in ts.get('timesheets', []):
            cid = entry.get('Client_ID')
            if not cid:
                continue
            if cid not in summary:
                summary[cid] = {
                    'client_id': cid,
                    'client_name': clients.get(cid, {}).get('Full_Name', cid),
                    'total_hours': 0,
                    'total_revenue': 0,
                    'max_monthly': clients.get(cid, {}).get('Max_Amount_Month', 0),
                    'funding_source': clients.get(cid, {}).get('Funding_Source', ''),
                }
            h = entry.get('Hours_Worked') or 0
            r = entry.get('Revenue') or 0
            summary[cid]['total_hours'] += h
            summary[cid]['total_revenue'] += r

        # Check FSW limit alerts
        for cid, s in summary.items():
            max_amt = s['max_monthly'] or 0
            used = s['total_revenue']
            s['fsw_used_pct'] = round((used / max_amt * 100), 1) if max_amt else 0
            s['fsw_alert'] = s['fsw_used_pct'] >= 80

        return {
            'month': month, 'year': year,
            'summaries': list(summary.values()),
            'total_revenue': sum(s['total_revenue'] for s in summary.values()),
            'total_hours': sum(s['total_hours'] for s in summary.values()),
        }
    except Exception as e:
        return {'error': str(e)}

def check_missing_data():
    """Check for missing data in clients and employees"""
    issues = []
    try:
        for c in get_clients().get('clients', []):
            if not c.get('Hourly_Charge_Rate_1'):
                issues.append(f"Client {c['Client_ID']} {c.get('Full_Name','')} missing hourly rate")
            if not c.get('Funding_Source'):
                issues.append(f"Client {c['Client_ID']} {c.get('Full_Name','')} missing funding source")
        for e in get_employees().get('employees', []):
            if not e.get('Hourly_Pay_Rate'):
                issues.append(f"Employee {e['Employee_ID']} {e.get('Full_Name','')} missing pay rate")
            if not e.get('Email'):
                issues.append(f"Employee {e['Employee_ID']} {e.get('Full_Name','')} missing email")
        return {'issues': issues, 'count': len(issues)}
    except Exception as ex:
        return {'error': str(ex)}

if __name__ == '__main__':
    print('=== PROAS Tools Test ===')
    print('\nClients:', get_clients())
    print('\nEmployees:', get_employees())
    print('\nMonthly Summary:', get_monthly_summary())
    print('\nMissing Data:', check_missing_data())
